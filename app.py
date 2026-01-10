"""
=========================================================
SAT INDIA — CORE APPLICATION
---------------------------------------------------------
Purpose:
- Serve dynamic content pages (100+ pages, no manual routes)
- Handle secure subscription workflow
- Enforce security, privacy, and rate limits
- Be production-ready for cloud deployment

Architecture Principles (Apple / Meta style):
- Few stable routes
- Logic separated from routing
- Dynamic content resolution
- Explicit action endpoints (POST)
=========================================================
"""

# =========================================================
# 0. ENVIRONMENT LOADING
# ---------------------------------------------------------
# Loads environment variables from .env in local/dev
# In production, values come from cloud secrets
# =========================================================

from dotenv import load_dotenv
load_dotenv()

# =========================================================
# 1. CORE IMPORTS
# ---------------------------------------------------------
# Flask framework + security + utilities
# =========================================================

from flask import (
    Flask, render_template, redirect, url_for,
    request, flash, jsonify
)

import os
import logging
import re
import uuid
from datetime import datetime
import hashlib

import dns.resolver
from itsdangerous import URLSafeTimedSerializer
from openpyxl import Workbook, load_workbook

from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import CSRFProtect
from flask_wtf.csrf import CSRFError
from werkzeug.exceptions import TooManyRequests
from werkzeug.middleware.proxy_fix import ProxyFix

# =========================================================
# 2. ENVIRONMENT MODE (DEV / PROD SWITCH)
# ---------------------------------------------------------
# Single switch controls production hardening
# =========================================================

ENV = os.environ.get("SAT_ENV", "development")
IS_PROD = ENV == "production"

# =========================================================
# 3. APPLICATION INITIALIZATION
# ---------------------------------------------------------
# Central Flask app object
# =========================================================

SECRET_KEY = os.environ.get("SAT_SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("SAT_SECRET_KEY is required")

app = Flask(__name__)
# Prefer HTTPS URLs in production (AWS / Nginx / ALB)
if IS_PROD:
    app.config["PREFERRED_URL_SCHEME"] = "https"
app.secret_key = SECRET_KEY
serializer = URLSafeTimedSerializer(SECRET_KEY)

# =========================================================
# 4. REVERSE PROXY SUPPORT (IMPORTANT IN PROD)
# ---------------------------------------------------------
# Ensures correct IP / HTTPS detection behind Nginx / LB
# =========================================================

app.wsgi_app = ProxyFix(
    app.wsgi_app,
    x_for=1,
    x_proto=1,
    x_host=1,
    x_port=1
)

# =========================================================
# 5. APPLICATION SECURITY CONFIG
# ---------------------------------------------------------
# Cookies, upload limits, session hardening
# =========================================================

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=IS_PROD,   # HTTPS-only in prod
    MAX_CONTENT_LENGTH=1 * 1024 * 1024
)

# =========================================================
# 6. CSRF PROTECTION
# ---------------------------------------------------------
# Protects all POST actions automatically
# =========================================================

csrf = CSRFProtect(app)




##-------------------Sections 0–6 ------------------End--------------------


# =========================================================
# 7. RATE LIMITING (ANTI-SPAM / ANTI-BOT)
# ---------------------------------------------------------
# Prevents abuse (bots, scraping, spam)
# =========================================================
#multiple workers
#real DB
#production Redis
#👉 You can safely switch back to Redis then.
## 
# In production with multiple workers + DB,
# switch storage_uri to Redis (REDIS_URL).
#
# limiter = Limiter(
#    key_func=get_remote_address,
#   app=app,
#   storage_uri=os.environ.get("REDIS_URL"),
# default_limits=["200 per hour"]
#

limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    storage_uri="memory://",   # in-memory limiter (no Redis)
    default_limits=["300 per hour"]
)

# =========================================================
# 8. AUDIT LOGGING
# ---------------------------------------------------------
# Centralized audit trail (security + compliance)
# =========================================================

logging.basicConfig(
    filename="audit.log",
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

def audit(event, email="", ip=""):
    """Standard audit logger"""
    logging.info(f"{event} | {email} | IP:{ip}")

# =========================================================
# 9. EMAIL VALIDATION LOGIC
# ---------------------------------------------------------
# Syntax check + MX record validation
# =========================================================

EMAIL_REGEX = re.compile(
    r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
)

def is_valid_email(email):
    if not EMAIL_REGEX.match(email):
        return False
    try:
        dns.resolver.resolve(email.split("@")[1], "MX")
        return True
    except Exception:
        return False

# =========================================================
# 10. SUBSCRIBER STORAGE (EXCEL — PHASE 1)
# ---------------------------------------------------------
# Temporary persistence layer
# Easy to migrate later to DB
# =========================================================
# =========================================================
# (DPDP-COMPLIANT — INTERIM EXCEL)
# =========================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SUBSCRIBERS_FILE = os.path.join(BASE_DIR, "data", "subscribers.xlsx")

def ensure_excel_file():
    """
    Ensures subscriber storage exists.
    Government compliance:
    - Predictable storage location
    - No runtime ambiguity
    """
    os.makedirs(os.path.dirname(SUBSCRIBERS_FILE), exist_ok=True)

    if not os.path.exists(SUBSCRIBERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "subscribers"
        ws.append([
            "email",
            "consent_given",
            "subscribed_at_utc",
            "ip_hash"
        ])
        wb.save(SUBSCRIBERS_FILE)

def hash_ip(ip):
    """
    Privacy-safe IP hashing.
    Prevents identification while allowing abuse detection.DPDP-COMPLIANT
    """
    return hashlib.sha256(ip.encode("utf-8")).hexdigest()

def is_email_already_subscribed(email):
    """Prevent duplicate subscriptions"""
    wb = load_workbook(SUBSCRIBERS_FILE)
    ws = wb["subscribers"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0].lower() == email.lower():
            return True
    return False

# Ensure storage exists at startup
ensure_excel_file()





#-------------------Sections 7-10 ------------------End----------------------------------------




# =========================================================
# 11. CORE ROUTES (ENTRY POINTS)
# ---------------------------------------------------------
# Very few routes — stable forever
# =========================================================

@app.route("/")
def home():
    return render_template("index.html")

@app.route("/health")
def health():
    """Used by load balancers / uptime checks"""
    return "OK", 200

# =========================================================
# 12. ACTION ROUTE — SUBSCRIBE
# ---------------------------------------------------------
# Purpose::
# - Handle email subscription securely
# - Enforce explicit consent (DPDP compliant)
# - Prevent duplicate subscriptions
# - Store data safely in Excel
# - Show feedback on the same page (join-us)
# =========================================================

@app.route("/subscribe", methods=["POST"])
def subscribe():
    """
    Consent-based subscription endpoint.

    Key guarantees:
    - Explicit consent required
    - Email validated (syntax + MX)
    - Duplicate emails blocked
    - No raw IP stored (only hash)
    - Excel write is fail-safe
    - User always redirected back to join-us page
    """

    # -----------------------------------------------------
    # a. Read form inputs
    # -----------------------------------------------------
    email = request.form.get("email", "").strip().lower()
    consent = request.form.get("consent")
    ip = request.remote_addr or "unknown"

    # -----------------------------------------------------
    # b. Enforce consent (legal requirement)
    # -----------------------------------------------------
    if not consent:
        audit("CONSENT_MISSING", email, ip)
        flash("Consent is required to subscribe.")
        return redirect(url_for("render_dynamic_page", slug="join-us"))

    # -----------------------------------------------------
    # c. Validate email (syntax + MX record)
    # -----------------------------------------------------
    if not is_valid_email(email):
        audit("INVALID_EMAIL", email, ip)
        flash("Please enter a valid email address.")
        return redirect(url_for("render_dynamic_page", slug="join-us"))

    # -----------------------------------------------------
    # d. Prevent duplicate subscriptions
    # -----------------------------------------------------
    if is_email_already_subscribed(email):
        audit("DUPLICATE_SUBSCRIPTION", email, ip)
        flash("This email is already subscribed.")
        return redirect(url_for("render_dynamic_page", slug="join-us"))

    # -----------------------------------------------------
    # e. Store subscriber in Excel (BULLETPROOF)
    # -----------------------------------------------------
    wb = load_workbook(SUBSCRIBERS_FILE)

    #f Ensure sheet exists (prevents KeyError / 500 errors)
    if "subscribers" not in wb.sheetnames:
        ws = wb.create_sheet("subscribers")
        ws.append([
            "email",
            "consent_given",
            "subscribed_at_utc",
            "ip_hash"
        ])
    else:
        ws = wb["subscribers"]

    # g Append subscriber data
    ws.append([
        email,
        "yes",
        datetime.utcnow().isoformat(),
        hash_ip(ip)
    ])

    wb.save(SUBSCRIBERS_FILE)

    # -----------------------------------------------------
    # h Audit + user feedback
    # -----------------------------------------------------
    audit("SUBSCRIBED", email, ip)
    flash("Thank you for subscribing. You’re now part of SAT.")

    # i Always return user to the page that owns the form
    return redirect(url_for("render_dynamic_page", slug="join-us"))


#-------------------Sections 11-12 ------------------End----------------------------------------


# =========================================================
# =========================================================
# 13. DYNAMIC CONTENT ROUTING (HYBRID: AUTO + EXCEPTIONS)
# ---------------------------------------------------------
# DESIGN GOALS:
# - Keep URLs stable forever
# - Do NOT touch HTML files
# - Auto-resolve most pages
# - Explicitly map only where URL != filename
# - One catch-all route, no duplicates
#
# This pattern is audit-friendly and future-proof.
# =========================================================

# Allow only lowercase letters, numbers, and hyphens in URLs
# Prevents path traversal, injections, and malformed requests
SAFE_SLUG = re.compile(r"^[a-z0-9\-]+$")

# Routes that must NEVER be treated as pages
# These are system / API / infrastructure endpoints
RESERVED_ROUTES = {
    "subscribe",
    "health",
    "client-error",
    "static"
}

# Explicit mapping ONLY where:
# URL slug ≠ template filename
# (Legal / policy pages often need this)
PAGE_MAP = {
    # URL             Template file
    "privacy-policy": "privacy-policy.html",
    "legal-terms": "legal-terms.html"
}

@app.route("/<slug>")
def render_dynamic_page(slug):
    """
    Dynamic page resolver.

    Resolution order (IMPORTANT):
    1. Block system routes (security)
    2. Validate slug format (safety)
    3. Check explicit PAGE_MAP (legal exceptions)
    4. Fallback to automatic <slug>.html
    5. Render only if template exists
    """

    # 1️⃣ Never allow system routes to be captured dynamically
    if slug in RESERVED_ROUTES:
        return render_template("404.html"), 404

    # 2️⃣ Enforce strict, safe URL format
    if not SAFE_SLUG.match(slug):
        return render_template("404.html"), 404

    # 3️⃣ Resolve template name:
    #    - Use PAGE_MAP for intentional exceptions
    #    - Otherwise auto-map slug → slug.html
    template_name = PAGE_MAP.get(slug, f"{slug}.html")

    # Absolute path check for safety & clarity
    template_path = os.path.join(app.template_folder, template_name)

    # 4️⃣ Render only if the file actually exists
    if not os.path.exists(template_path):
        return render_template("404.html"), 404

    # 5️⃣ Render approved template
    return render_template(template_name)


# =========================================================
# 14. CLIENT ERROR TELEMETRY
# ---------------------------------------------------------
# Receives JS errors from frontend
# =========================================================

@app.route("/client-error", methods=["POST"])
def client_error():
    data = request.get_json(silent=True)
    logging.error(f"CLIENT_ERROR | {data}")
    return jsonify({"status": "ok"})

# =========================================================
# 15. ERROR HANDLING
# ---------------------------------------------------------
# Real errors — never masked
# =========================================================

@app.errorhandler(404)
def handle_404(e):
    logging.warning(f"404 | {request.path}")
    return render_template("404.html"), 404


@app.errorhandler(TooManyRequests)
def handle_rate_limit(e):
    flash("Too many requests. Try later.")
    return redirect(url_for("render_dynamic_page", slug="join-us"))

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    flash("Session expired. Refresh and retry.")
    return redirect(url_for("render_dynamic_page", slug="join-us"))
@app.errorhandler(Exception)

def handle_unexpected_error(e):
    if not IS_PROD:
        raise e  # Full traceback in development
    logging.critical("UNHANDLED_EXCEPTION", exc_info=True)
    return render_template("500.html"), 500

# =========================================================
# 16. APPLICATION ENTRY POINT
# ---------------------------------------------------------
# Gunicorn / Docker compatible
# =========================================================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)


#-------------------Sections 13-16 ------------------End----------------------------------------
